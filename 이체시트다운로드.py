from auth import get_credentials
from googleapiclient.discovery import build
import logging
from openpyxl import load_workbook
import os
from openpyxl.utils import get_column_letter
import webbrowser  # 웹브라우저 모듈 추가
from pathlib import Path  

def get_data_from_sheets(spreadsheet_id, range_name):
    """Google Sheets에서 데이터를 가져오는 함수"""
    logging.info(f"스프레드시트에서 데이터 가져오기 시작: {range_name}")
    try:
        creds = get_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        sheet = service.spreadsheets()
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=range_name
        ).execute()
        
        values = result.get('values', [])
        
        # 데이터 전처리: 쉼표 제거
        processed_values = []
        for row in values:
            processed_row = []
            for cell in row:
                # 문자열인 경우에만 쉼표 제거
                if isinstance(cell, str):
                    cell = cell.replace(',', '')
                processed_row.append(cell)
            processed_values.append(processed_row)
        
        logging.info(f"데이터 가져오기 성공: {len(processed_values)} rows")
        return processed_values
    
    except Exception as e:
        logging.error(f"데이터 가져오기 실패: {str(e)}")
        raise

def update_excel_file(excel_path, sheet_ranges):
    """엑셀 파일 업데이트 함수"""
    logging.info("엑셀 파일 업데이트 시작")
    
    SPREADSHEET_ID = '1NOP5_s0gNUCWaGIgMo5WZmtqBbok_5a4XdpNVwu8n5c'
    processed_data = []  # 처리된 데이터를 저장할 리스트
    total_rows = 0
    
    try:
        # 엑셀 파일이 없으면 생성
        if not os.path.exists(excel_path):
            from openpyxl import Workbook
            wb = Workbook()
            wb.save(excel_path)
        
        # 엑셀 파일 로드
        wb = load_workbook(excel_path)
        
        # 각 시트 범위에 대해 처리
        for sheet_num, (source_range, target_sheet, target_range) in enumerate(sheet_ranges, 1):
            logging.info(f"시트 {sheet_num} 처리 중...")
            
            data = get_data_from_sheets(SPREADSHEET_ID, f"시트1!{source_range}")
            total_rows += len(data)
            processed_data.append({
                "시트번호": sheet_num,
                "시트명": target_sheet,
                "데이터": data
            })
            
            # 시트가 없으면 생성
            if target_sheet not in wb.sheetnames:
                wb.create_sheet(target_sheet)
            ws = wb[target_sheet]
            
            # 기존 데이터 삭제 (헤더 제외)
            target_start_row = int(target_range.split(':')[0][1:])
            target_end_row = int(target_range.split(':')[1][1:])
            target_start_col = target_range.split(':')[0][0]
            target_end_col = target_range.split(':')[1][0]
            
            # 헤더를 제외한 기존 데이터 삭제
            for row in range(target_start_row + 1, target_end_row + 1):
                for col in range(ord(target_start_col) - ord('A') + 1, 
                               ord(target_end_col) - ord('A') + 2):
                    ws[f"{get_column_letter(col)}{row}"] = None
            
            # 새 데이터 입력
            for i, row_data in enumerate(data):
                for j, value in enumerate(row_data):
                    col = get_column_letter(ord(target_start_col) - ord('A') + 1 + j)
                    ws[f"{col}{target_start_row + i}"] = value
        
        # 변경사항 저장
        wb.save(excel_path)
        logging.info("엑셀 파일 업데이트 완료")
        return total_rows, processed_data
        
    except Exception as e:
        logging.error(f"엑셀 파일 업데이트 실패: {str(e)}")
        raise

def main():
    # 로깅 설정
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    # 사용자에게 스프레드시트를 열지 물어보기
    open_sheets = input("입금요청 내역과 현황판을 함께 열까요? (y/n): ").strip().lower()
    
    if open_sheets == 'y':
        # 두 개의 스프레드시트 링크 열기
        webbrowser.open("https://docs.google.com/spreadsheets/d/1NOP5_s0gNUCWaGIgMo5WZmtqBbok_5a4XdpNVwu8n5c/edit?gid=0#gid=0")
        webbrowser.open("https://docs.google.com/spreadsheets/d/1CK2UXTy7HKjBe2T0ovm5hfzAAKZxZAR_ev3cbTPOMPs/edit?gid=1565864271#gid=1565864271")
    
    # 엑셀 파일 경로
    current_dir = Path(__file__).parent
    excel_path = str(current_dir / "이체정보.xlsx")
    
    # 시트 범위 매핑 (소스범위, 대상시트명, 대상범위)
    sheet_ranges = [
        ('E2:K11', 'Sheet1', 'D2:J11'),
        ('E12:K21', 'Sheet2', 'D2:J11'),
        ('E22:K31', 'Sheet3', 'D2:J11'),
        ('E32:K41', 'Sheet4', 'D2:J11'),
        ('E42:K51', 'Sheet5', 'D2:J11'),
        ('E52:K61', 'Sheet6', 'D2:J11'),
        ('E62:K71', 'Sheet7', 'D2:J11'),
        ('E72:K81', 'Sheet8', 'D2:J11'),
        ('E82:K91', 'Sheet9', 'D2:J11'),
        ('E92:K101', 'Sheet10', 'D2:J11')
    ]
    
    try:
        total_rows, processed_data = update_excel_file(excel_path, sheet_ranges)
        
        # 데이터가 있는 시트만 필터링
        valid_data = [data for data in processed_data if data['데이터']]
        
        print("\n=== 상세 데이터 ===")
        
        for sheet_data in valid_data:
            if sheet_data['데이터']:  # 데이터가 있는 경우만 출력
                print(f"\n[시트 {sheet_data['시트번호']} - {sheet_data['시트명']}]")
                for row in sheet_data['데이터']:
                    print(f"  {' | '.join(str(cell) for cell in row)}")
        
        print("\n=== 작업 완료 보고서 ===")
        print(f"총 처리된 시트 수: {len(valid_data)}")
        print(f"총 처리된 데이터 행 수: {sum(len(data['데이터']) for data in valid_data)}")
        print("\n모든 작업이 성공적으로 완료되었습니다.")
    except Exception as e:
        print(f"작업 중 오류가 발생했습니다: {str(e)}")

if __name__ == '__main__':
    main()