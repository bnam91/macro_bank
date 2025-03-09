import openpyxl
from pathlib import Path

def organize_excel_sheets(file_path):
    try:
        # 엑셀 파일 로드
        workbook = openpyxl.load_workbook(file_path)
        
        while True:
            # 워크북을 다시 로드하여 현재 상태 확인
            workbook = openpyxl.load_workbook(file_path)
            current_sheets = workbook.sheetnames
            print(f"\n현재 맨 앞에 있는 시트명은 '{current_sheets[0]}'입니다.")
            
            # 사용자 입력 받기
            user_input = input("어떤 시트를 맨앞으로 가져올까요? (숫자 또는 시트명 입력, 'clean'으로 정리, 'exit'로 종료): ")
            
            if user_input.lower() == 'exit':
                break
                
            if user_input.lower() == 'clean':
                # Sheet1부터 Sheet10까지 순서대로 정렬
                standard_sheets = sorted([s for s in current_sheets if s.startswith('Sheet') and s[5:].isdigit()], 
                                      key=lambda x: int(x[5:]))  # Sheet 뒤의 숫자로 정렬
                other_sheets = [sheet for sheet in current_sheets if sheet not in standard_sheets]
                
                # 시트 순서 재배열
                sheet_order = standard_sheets + other_sheets
                workbook._sheets = [workbook[sheet_name] for sheet_name in sheet_order]
                print("시트를 Sheet1-10 순서로 정리했습니다.")
                
            else:
                # 숫자나 시트명으로 입력받은 경우
                target_sheet = None
                
                # 숫자로 입력받은 경우
                if user_input.isdigit():
                    sheet_name = f"Sheet{user_input}"
                    if sheet_name in current_sheets:
                        target_sheet = sheet_name
                        
                # 시트명으로 입력받은 경우
                elif user_input in current_sheets:
                    target_sheet = user_input
                    
                if target_sheet:
                    # 시트 순서 변경
                    sheets = workbook._sheets
                    
                    # 먼저 Sheet1을 맨 뒤로 이동 (target이 Sheet1이 아닌 경우에만)
                    if target_sheet != "Sheet1" and "Sheet1" in current_sheets:
                        sheet1_idx = current_sheets.index("Sheet1")
                        sheets.append(sheets.pop(sheet1_idx))
                        
                    # 그 다음 선택한 시트를 맨 앞으로 이동
                    target_idx = workbook.sheetnames.index(target_sheet)  # 현재 시트 위치 다시 확인
                    sheet_to_move = sheets.pop(target_idx)
                    sheets.insert(0, sheet_to_move)
                    
                    print(f"'{target_sheet}'를 맨 앞으로 이동했습니다.")
                else:
                    print(f"해당 시트를 찾을 수 없습니다.")
                    print("현재 시트 목록:", current_sheets)
            
            # 변경사항 저장
            workbook.save(file_path)
            
        workbook.close()
        
    except Exception as e:
        print(f"에러가 발생했습니다: {str(e)}")
        print("현재 시트 목록:", workbook.sheetnames)

# 파일 경로 설정 및 실행
current_dir = Path(__file__).parent  # 현재 스크립트 파일의 디렉토리
file_path = current_dir / "이체정보.xlsx"  # 상대 경로로 파일 지정
organize_excel_sheets(str(file_path))  # Path 객체를 문자열로 변환