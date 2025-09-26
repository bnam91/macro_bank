import subprocess
import pyautogui
import time as tm
import cv2
import numpy as np
import mss
import os   
import pandas as pd
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup

# 현재 스크립트의 절대 경로를 얻고, 그 디렉토리로 작업 디렉토리 변경
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# 이미지 파일 경로를 src 폴더에서 찾도록 설정
img_dir = os.path.join(script_dir, "src")
print("이미지 디렉토리:", img_dir)

# 사용자에게 다계좌이체진행 여부를 묻기
user_input = input("🟠다계좌이체진행(자동)을 함께 진행할까요? (y/n): ")
auto_transfer = user_input.lower() == 'y'

# 이미지 파일들을 src 폴더에서 찾도록 설정
COORDS = {
    "localdisktab": "localdisk.png",
    "outdisktab": "seagate_bt.png",
    "commonlogin": "localdisk_loginbt.png",
    "send_tab": "send_tab.png",
    "multisend": "multisend_bt.png",
    "password4_input": "password4_input.png",
    # 키보드 버튼 추가
    "shift_bt": "shift_bt.png",
    "golbang_bt": "golbang_bt.png",
    "shift2_bt": "shift2_bt.png",
    "key_g": "key_g.png",
    "key_u": "key_u.png",
    "key_s": "key_s.png",
    "key_q": "key_q.png",
    "key_l": "key_l.png",
    "key_s": "key_s.png",
    "key_1": "key_1.png",
    "key_2": "key_2.png",
    "key_0": "key_0.png",
    "key_enter": "key_enter.png",
}

# 이미지 경로를 가져오는 함수 추가
def get_image_path(image_name):
    # 이미지 경로를 유니코드로 변환하여 처리
    path = os.path.join(img_dir, image_name)
    return os.path.normpath(path)

# 모니터 설정 부분을 동적으로 변경
def get_monitor_configs():
    with mss.mss() as sct:
        monitors = []
        for i, monitor in enumerate(sct.monitors[1:], 1):  # monitors[0]은 전체 화면이므로 제외
            monitors.append({
                'top': monitor['top'],
                'left': monitor['left'],
                'width': monitor['width'],
                'height': monitor['height'],
                'number': i
            })
            print(f"모니터 {i} 감지됨: {monitors[-1]}")
        return monitors

# 전역 변수로 모니터 설정
MONITORS = get_monitor_configs()

def capture_screen(region):
    with mss.mss() as sct:
        screenshot = sct.grab(region)
        img = np.array(screenshot)
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)
        return img                      

def match_template(screen, template_path):
    # 전체 경로로 이미지 로드
    full_path = get_image_path(template_path)
    # cv2.imread 대신 numpy를 사용하여 이미지 로드
    template = cv2.imdecode(np.fromfile(full_path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if template is None:
        raise ValueError(f"이미지를 찾을 수 없습니다: {full_path}")

    result = cv2.matchTemplate(screen, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)
    if max_val >= 0.98:  # 일치율이 90% 이상인 경우
        return (max_loc[0] + template.shape[1] // 2, max_loc[1] + template.shape[0] // 2)
    return None

def locate_image_on_monitors(image_path):
    """모든 감지된 모니터에서 이미지를 찾습니다."""
    for monitor in MONITORS:
        location = locate_image_on_monitor(image_path, monitor)
        if location:
            return location
    return None
    
def locate_image_on_monitor(image_path, monitor):
    """단일 모니터에서 이미지를 찾습니다."""
    screen = capture_screen(monitor)
    location = match_template(screen, image_path)
    if location:
        location = (location[0] + monitor['left'], location[1] + monitor['top'])
    return location

def locate_and_click(image_name):
    image_path = COORDS[image_name]  # 이미지 파일명 가져오기
    print(f"{image_path}을(를) 찾고 클릭합니다.")
    
    # 최대 20회 재시도
    max_attempts = 20
    for attempt in range(max_attempts):
        location = locate_image_on_monitors(image_path)
        if location:
            print(f"{image_path}의 위치를 찾았습니다: {location}, 클릭합니다.")
            pyautogui.click(location)
            tm.sleep(1)
            return True
        else:
            if attempt < max_attempts - 1:  # 마지막 시도가 아니면
                print(f"{image_path}의 위치를 찾을 수 없습니다. 1초 후 재시도합니다. ({attempt+1}/{max_attempts})")
                tm.sleep(1)  # 1초 대기 후 재시도
            else:
                print(f"{image_path}의 위치를 찾을 수 없습니다. 최대 시도 횟수 초과.")
                return False

def locate_and_click_with_retry(image_name, max_retries=30, retry_interval=1):
    """이미지를 찾고 클릭하는 함수 (재시도 로직 포함)"""
    image_path = COORDS[image_name]  # 이미지 파일명 가져오기
    print(f"{image_path}을(를) 찾고 클릭합니다.")
    
    for attempt in range(max_retries):
        location = locate_image_on_monitors(image_path)
        if location:
            print(f"{image_path}의 위치를 찾았습니다: {location}, 클릭합니다. (시도 {attempt+1}/{max_retries})")
            pyautogui.click(location)
            tm.sleep(1)
            return True
        else:
            print(f"{image_path}의 위치를 찾을 수 없습니다. (시도 {attempt+1}/{max_retries})")
            if attempt < max_retries - 1:  # 마지막 시도가 아니면
                tm.sleep(retry_interval)  # 1초 대기 후 재시도
    
    print(f"{image_path}을(를) {max_retries}초 동안 찾을 수 없었습니다.")
    return False

# Selenium 설정 및 웹페이지 열기
url = 'https://www.kebhana.com/common/login.do'
options = Options()
options.add_experimental_option("detach", True)
options.add_argument("--disable-blink-features=AutomationControlled")

driver = webdriver.Chrome(options=options)
driver.get(url)

try:
    # 최대 30초 동안 대기하되, 요소가 나타나면 즉시 진행
    wait = WebDriverWait(driver, 30)
    
    # 먼저 "공동/금융인증서 로그인" 메뉴 클릭
    try:
        # 공동/금융인증서 로그인 메뉴 찾기 및 클릭
        cert_menu_element = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), '공동/금융인증서 로그인')]")))
        cert_menu_element.click()
        print("공동/금융인증서 로그인 메뉴 클릭 성공")
        tm.sleep(3)  # 3초 대기
    except Exception as e:
        print(f"공동/금융인증서 로그인 메뉴 클릭 실패: {e}")
    
    # 기존 공동인증서 로그인 버튼 클릭
    cert_login_element = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#certLogin")))
    print(cert_login_element.text)
    
    # 클릭 시도와 재시도 로직
    max_attempts = 60  # 최대 시도 횟수
    for attempt in range(max_attempts):
        try:
            # JavaScript로 클릭 시도 (더 안정적)
            driver.execute_script("arguments[0].click();", cert_login_element)
            print("공동인증서 로그인 버튼 클릭 성공")
            break  # 성공하면 반복문 종료
        except Exception as e:
            if attempt < max_attempts - 1:  # 마지막 시도가 아니면
                print(f"클릭 실패, 1초 후 재시도합니다. ({attempt+1}/{max_attempts})")
                tm.sleep(1)  # 1초 대기 후 재시도
            else:
                print(f"최대 시도 횟수 초과: {e}")
                raise  # 재시도 모두 실패 시 예외 발생
except TimeoutException:
    print("certLogin 버튼을 찾을 수 없습니다.")
    driver.quit()

tm.sleep(10)  # 이후 작업을 위한 대기 시간은 유지

# PyAutoGUI를 이용한 인증서 로그인
def login_with_certificate():
    # localdisk.png를 30초간 1초 간격으로 찾기
    if not locate_and_click_with_retry("localdisktab", max_retries=30, retry_interval=1):
        print("localdisk.png를 찾을 수 없어 로그인을 중단합니다.")
        return
    
    tm.sleep(3)
    locate_and_click("outdisktab")     # 외장하드선택
    tm.sleep(3)     
    
    for _ in range(6):   # 탭 6번
        pyautogui.press('tab')
        tm.sleep(0.15)
    
    # 화면 전환을 위한 대기 시간 추가
    tm.sleep(2)  # 2초 대기0
    
    # shift_bt 이미지가 있는지 확인
    shift_location = locate_image_on_monitors(COORDS["shift_bt"])
    
    if shift_location:  # shift_bt가 있으면 (None이 아니면)
        # shift_bt가 있으면 마우스 클릭으로 입력
        print("마우스 클릭으로 비밀번호 입력")
        keyboard_sequence = [
            "shift_bt",
            "golbang_bt",
            "shift2_bt",
            "key_g",
            "key_u",
            "key_s",
            "key_q",
            "key_l",
            "key_s",
            "key_1",
            "key_2",
            "key_0",
            "key_enter"
        ]
        
        for key in keyboard_sequence:
            locate_and_click(key)
            tm.sleep(0.1)  # 각 키 입력 사이에 0.1초 대기
    else:
        # shift_bt가 없으면 키보드로 입력
        print("키보드로 비밀번호 입력")
        password = "@gusqls120"
        for char in password:
            pyautogui.press(char)
            tm.sleep(0.07)
    
    locate_and_click("commonlogin")    # 공동로그인 버튼 클릭

login_with_certificate()
tm.sleep(8)

# 프레임 전환
frame_id = "hanaMainframe"  # 프레임 ID를 여기에 입력하세요
driver.switch_to.frame(frame_id)

# 이체 메뉴 직접 클릭 (send_tab.png 대신 HTML 요소 클릭)
try:
    # XPath를 사용하여 "이체" 링크 찾기 (585번 라인 부근의 요소)
    transfer_link = driver.find_element(By.XPATH, "//a[@title='이체' and text()='이체']")
    print("'이체' 메뉴 링크를 찾았습니다.")
    
    # 링크 클릭
    transfer_link.click()
    print("'이체' 메뉴 링크를 클릭했습니다.")
    tm.sleep(1)
except Exception as e:
    print(f"'이체' 메뉴 링크 클릭 중 오류 발생: {e}")
    # 실패했을 경우 기존 방식 시도
    print("기존 방식으로 이체 탭 클릭을 시도합니다.")                       
    
    locate_and_click("send_tab")

tm.sleep(1)

# 다계좌 이체 버튼 클릭
for _ in range(6):   # 탭 3번
    pyautogui.press('tab')

pyautogui.press('enter')   
tm.sleep(3)

def scroll_down(amount):
    pyautogui.scroll(-amount)  # 스크롤 다운

def scroll_up(amount):
    pyautogui.scroll(amount)  # 스크롤 업

# 예시: 전체 페이지를 스크롤 다운하고 절반만 다시 스크롤 업
scroll_amount = 1000  # 페이지의 길이에 맞게 조정
scroll_down(scroll_amount)
tm.sleep(1)  # 잠시 대기
scroll_up(scroll_amount // 2)

# 은행 선택을 위한 맵핑
bank_options = {
    "하나은행": "081",
    "경남은행": "039",
    "광주은행": "034",
    "국민은행": "004",
    "기업은행": "003",
    "농협": "011",
    "iM뱅크(대구)": "031",
    "도이치뱅크": "055",
    "부산은행": "032",
    "산업은행": "002",
    "저축은행": "050",
    "새마을금고": "045",
    "수협은행": "007",
    "신협": "048",
    "신한은행": "088",
    "우리은행": "020",
    "우체국": "071",
    "전북은행": "037",
    "제주은행": "035",
    "카카오뱅크": "090",
    "케이뱅크": "089",
    "한국씨티은행": "027", 
    "BOA": "060",
    "HSBC": "054",
    "JP모간": "057",
    "SC제일은행": "023",
    "하나증권": "270",
    "교보증권": "261",
    "대신증권": "267",
    "미래에셋증권": "238",
    "DB금융투자": "279",
    "유안타증권": "209",
    "메리츠증권": "287",
    "부국증권": "290",
    "삼성증권": "240",
    "신영증권": "291",
    "신한투자증권": "278",
    "NH투자증권": "247",
    "유진증권": "280",
    "키움증권": "264",
    "하이투자증권": "262",
    "한국투자": "243",
    "한화투자증권": "269",
    "KB증권": "218",
    "LS증권": "265",
    "현대차증권": "263",
    "케이프증권": "292",
    "SK증권": "266",
    "산림조합": "064",
    "중국공상은행": "062",
    "중국은행": "063",
    "중국건설은행": "067",
    "BNP파리바은행": "061",
    "한국포스증권": "294",
    "다올투자증권": "227",
    "BNK투자증권": "224",
    "카카오페이증권": "288",
    "IBK투자증권": "225",
    "토스증권": "271",
    "토스뱅크": "092",
    "상상인증권": "221"
}

# 은행명 표준화 함수
def standardize_bank_name(bank_name):
    bank_name = bank_name.lower().strip()
    if "sc제일" in bank_name or "제일" in bank_name:
        return "SC제일은행"
    elif "제일은행" in bank_name:
        return "SC제일은행"
    elif "하나" in bank_name:
        return "하나은행"
    elif "경남" in bank_name:
        return "경남은행"
    elif "광주" in bank_name:
        return "광주은행"
    elif "국민" in bank_name:
        return "국민은행"
    elif "기업" in bank_name:
        return "기업은행"
    elif "농협은행" in bank_name:
        return "농협"
    elif "NH농협" in bank_name:
        return "농협"
    elif "nh농협" in bank_name:
        return "농협"
    elif "농협/" in bank_name:
        return "농협"
    elif "대구" in bank_name:         
        return "iM뱅크(대구)"
    elif "im뱅크" in bank_name:
        return "iM뱅크(대구)"
    elif "대구은행" in bank_name:
        return "iM뱅크(대구)"
    elif "부산" in bank_name:
        return "부산은행"
    elif "새마을" in bank_name:
        return "새마을금고"
    elif "수협" in bank_name:
        return "수협은행"
    elif "신한" in bank_name:
        return "신한은행"
    elif "우리" in bank_name:
        return "우리은행"
    elif "전북" in bank_name:
        return "전북은행"
    elif "제주" in bank_name:
        return "제주은행"
    elif "카카오" in bank_name:
        return "카카오뱅크"
    elif "카뱅" in bank_name:
        return "카카오뱅크"
    elif "씨티" in bank_name:
        return "한국씨티은행"
    elif "토스" in bank_name:
        return "토스뱅크"
    elif "카카오페이증권" in bank_name:
        return "카카오페이증권"    
    elif "미래에셋대우" in bank_name:
        return "미래에셋증권"
    elif "미래에셋" in bank_name:
        return "미래에셋증권"
       
       
    # 필요한 경우 추가 은행명 표준화
    return bank_name



# 전처리 함수
def preprocess_account_info(account_info):
    # 은행명과 계좌번호 분리
    match = re.match(r'(\D+)\s*([\d\s\-]+)', account_info)
    if match:
        bank_name = standardize_bank_name(match.group(1).strip())
        account_number = re.sub(r'[\s\-]', '', match.group(2))
    else:
        bank_name = ""
        account_number = ""
    
    return bank_name, account_number

# 엑셀 파일 로드
excel_path = os.path.join(script_dir, "이체정보.xlsx")
df = pd.read_excel(excel_path)

# 전처리된 데이터를 저장할 리스트
processed_data = []

for index, row in df.iterrows():
    product_name = row.iloc[3]  # 제품명
    customer_name = row.iloc[4]  # 이름
    account_info = row.iloc[7]  # 은행+계좌번호
    amount = row.iloc[9]  # 금액

    # 계좌정보 전처리
    bank_name, account_number = preprocess_account_info(account_info)

    # 항목 1 : 은행
    # 항목 2 : 계좌번호             
    # 항목 3 : 이름.제품명
    # 항목 4 : 제품명
    name_product = f"{customer_name}{product_name}"
    processed_data.append((bank_name, account_number, name_product, product_name, amount))

# 전처리된 데이터 출력
for data in processed_data:
    print(f"은행: {data[0]}, 계좌번호: {data[1]}, 이름.제품명: {data[2]}, 제품명: {data[3]}, 금액: {data[4]}")

def type_text(element, text):
    element.clear()
    element.send_keys(text)

def type_number(text):
    for char in str(text):
        pyautogui.press(char)
        tm.sleep(0.005)

def input_transfer_info(data, index):
    bank, account_number, name_product, product_name, amount = data
    
    try:
        # 은행 선택
        bank_element = driver.find_element(By.ID, f"rcvBnkCd{index}")
        option_value = bank_options.get(bank, "")
        if option_value:
            option = bank_element.find_element(By.CSS_SELECTOR, f"option[value='{option_value}']")  
            option.click()
            print(f"{name_product} 은행 선택 성공: {bank}")
        else:
            print(f"{name_product} 은행을 찾을 수 없습니다: {bank}")
        
        # 계좌번호 입력
        account_element = driver.find_element(By.ID, f"rcvAcctNo{index}")
        account_element.click()
        tm.sleep(0.1)  # 첫 번째 문자 입력 전에 짧은 지연을 추가
        type_number(account_number)
        print(f"{name_product} 계좌번호 입력 성공: {account_number}")
        
        # 금액 입력
        amount_element = driver.find_element(By.ID, f"trnsAmt{index}")
        amount_element.click()
        tm.sleep(0.1)  # 첫 번째 문자 입력 전에 짧은 지연을 추가
        type_number(int(amount))
        print(f"{name_product} 금액 입력 성공: {amount}")
        
        # 이름.제품명 입력
        name_product_element = driver.find_element(By.ID, f"wdrwPsbkMarkCtt{index}")
        type_text(name_product_element, name_product)
        print(f"{name_product} 이름.제품명 입력 성공: {name_product}")
        
        # 제품명 입력
        product_name_element = driver.find_element(By.ID, f"rcvPsbkMarkCtt{index}")
        type_text(product_name_element, product_name)
        print(f"{name_product} 제품명 입력 성공: {product_name}")

    except Exception as e:
        print(f"오류 발생 {name_product}: {e}")

# paymAcctPw에 '5800' 입력
def enter_password():
    try:
        # 스크롤을 맨 위로 올려서 계좌비밀번호 필드로 이동
        driver.execute_script("window.scrollTo(0, 0)")
        tm.sleep(0.5)  # 스크롤 후 잠시 대기
        
        password_element = driver.find_element(By.ID, "paymAcctPw")
        password_element.click()
        tm.sleep(0.1)  # 첫 번째 문자 입력 전에 짧은 지연을 추가
        type_number("5800")
        print("비밀번호 입력 성공")
    except Exception as e:
        print(f"비밀번호 입력 오류: {e}")

# 다계좌이체진행 버튼 클릭
def click_transfer_button():
    try:
        # 1~2초 대기
        tm.sleep(2)
                                
        # 스크롤을 아래로 내려서 버튼 위치로 이동
        driver.execute_script("window.scrollBy(0, 500)")
        tm.sleep(0.5)  # 스크롤 후 잠시 대기
        
        # 다계좌이체진행 버튼 찾기 및 클릭
        transfer_button = driver.find_element(By.XPATH, "//a[contains(text(), '다계좌이체진행')]")
        transfer_button.click()
        print("다계좌이체진행 버튼 클릭 성공")
    except Exception as e:
        print(f"다계좌이체진행 버튼 클릭 오류: {e}")

# 보이스피싱 예방 팝업 처리 함수
def handle_voice_phishing_popup():
    try:
        # 최대 60초 동안 1초마다 팝업 확인
        for i in range(60):
            tm.sleep(1)  # 1초 대기
            print(f"팝업 감지 중... {i+1}초 경과")
            
            # 보이스피싱 예방 팝업 존재 여부 확인 - 두 가지 ID 모두 확인
            voice_phishing_popup1 = driver.find_elements(By.ID, "voicePhishingPopup1")
            voice_phishing_popup2 = driver.find_elements(By.ID, "lonFrdInfoPop")
            
            # 첫 번째 유형의 보이스피싱 팝업 확인
            if voice_phishing_popup1 and len(voice_phishing_popup1) > 0 and voice_phishing_popup1[0].is_displayed():
                print("보이스피싱 예방 팝업(voicePhishingPopup1) 감지됨")
                
                # '아니요' 버튼 클릭
                no_button = driver.find_element(By.XPATH, "//a[contains(@onclick, 'pbk.transfer.common.lonFrdInfoPopN()')]")
                no_button.click()
                print("'아니요' 버튼 클릭 성공")
                return True
                
            # 두 번째 유형의 보이스피싱 팝업 확인
            if voice_phishing_popup2 and len(voice_phishing_popup2) > 0 and voice_phishing_popup2[0].is_displayed():
                print("보이스피싱 예방 팝업(lonFrdInfoPop) 감지됨")
                
                # '아니요' 버튼 클릭
                no_button = driver.find_element(By.XPATH, "//a[contains(@onclick, 'pbk.transfer.common.lonFrdInfoPopN()')]")
                no_button.click()
                print("'아니요' 버튼 클릭 성공")
                return True
                
        print("60초 동안 보이스피싱 예방 팝업이 나타나지 않음")
        return False
    except Exception as e:
        print(f"팝업 처리 중 오류 발생: {e}")
        return False

# 최대 10개의 항목 입력
for index, data in enumerate(processed_data):
    if index >= 10:
        break
    input_transfer_info(data, index)
    tm.sleep(0.5)  # 각 세트 완료 후 잠시 대기

# 이체 정보 입력 후 비밀번호 입력
enter_password()

# 자동/수동에 따른 처리
if auto_transfer:
    # 다계좌이체진행 버튼 클릭
    click_transfer_button()

    # 보이스피싱 예방 팝업 처리
    handle_voice_phishing_popup()
    print("이체가 완료되었습니다.")     
else:
    print("이체가 취소되었습니다. 필요시 수동으로 다계좌이체진행 버튼을 클릭하세요.")

# 이체 작업 완료 후 시트 순서 변경 여부 묻기
print("\n" + "="*50)
print("이체 작업이 완료되었습니다!")
print("="*50)

# 무한 루프로 계속 진행
while True:
    sheet_organize_input = input("🟢시트 순서를 변경하시겠습니까? (y/n): ")
    if sheet_organize_input.lower() == 'y':
        print("시트 순서 변경을 시작합니다...")
    
    # 1. 시트순서변경.py의 기능을 여기에 통합
    import openpyxl
    from pathlib import Path
    
    def organize_excel_sheets(file_path):
        try:
            # 엑셀 파일 로드
            workbook = openpyxl.load_workbook(file_path)
            
            while True:
                # 워크북을 다시 로드하여 현재 상태 확인
                workbook = openpyxl.load_workbook(file_path)
                current_sheets = workbook.sheetnames
                print(f"\n현재 맨 앞에 있는 시트명은 👉 '{current_sheets[0]}'입니다.")
                
                # 사용자 입력 받기
                user_input = input("🟢어떤 시트를 맨앞으로 가져올까요? (숫자 또는 시트명 입력, 'clean'으로 정리, 'exit'로 종료): ")
                
                if user_input.lower() == 'exit':
                    break
                    
                if user_input.lower() == 'clean':
                    # Sheet1부터 Sheet10까지 순서대로 정렬
                    standard_sheets = sorted([s for s in current_sheets if s.startswith('Sheet') and s[5:].isdigit()], 
                                          key=lambda x: int(x[5:]))  # Sheet 뒤의 숫자로 정렬
                    other_sheets = [sheet for sheet in current_sheets if sheet not in standard_sheets]
                    
                    # 시트 순서 재배열
                    sheet_order = standard_sheets + other_sheets
                    workbook._sheets = [workbook[sheet_name] for sheet_name in sheet_order]
                    print("시트를 Sheet1-10 순서로 정리했습니다.")
                    
                else:
                    # 숫자나 시트명으로 입력받은 경우
                    target_sheet = None
                    
                    # 숫자로 입력받은 경우
                    if user_input.isdigit():
                        sheet_name = f"Sheet{user_input}"
                        if sheet_name in current_sheets:
                            target_sheet = sheet_name
                            
                    # 시트명으로 입력받은 경우
                    elif user_input in current_sheets:
                        target_sheet = user_input
                        
                    if target_sheet:
                        # 시트 순서 변경 - 1.시트순서변경.py와 동일한 방법
                        sheets = workbook._sheets
                        
                        # 먼저 Sheet1을 맨 뒤로 이동 (target이 Sheet1이 아닌 경우에만)
                        if target_sheet != "Sheet1" and "Sheet1" in current_sheets:
                            sheet1_idx = current_sheets.index("Sheet1")
                            sheets.append(sheets.pop(sheet1_idx))
                            
                        # 그 다음 선택한 시트를 맨 앞으로 이동
                        target_idx = workbook.sheetnames.index(target_sheet)  # 현재 시트 위치 다시 확인
                        sheet_to_move = sheets.pop(target_idx)
                        sheets.insert(0, sheet_to_move)
                        
                        print(f"'{target_sheet}'를 맨 앞으로 이동했습니다.")
                    else:
                        print(f"해당 시트를 찾을 수 없습니다.")
                        print("현재 시트 목록:", current_sheets)
                
                # 변경사항 저장
                workbook.save(file_path)
                
            workbook.close()
            
        except Exception as e:
            print(f"에러가 발생했습니다: {str(e)}")
            print("현재 시트 목록:", workbook.sheetnames)
    
    # 시트 순서 변경 실행
    current_dir = Path(__file__).parent  # 현재 스크립트 파일의 디렉토리
    file_path = current_dir / "이체정보.xlsx"  # 상대 경로로 파일 지정
    organize_excel_sheets(str(file_path))  # Path 객체를 문자열로 변환
    
    print("시트 순서 변경이 완료되었습니다.")
    
    # 시트 변경 후 이체 진행 여부 묻기
    retry_transfer_input = input("🟢시트를 변경했으니 다시 이체를 진행할까요? (y/n): ")
    if retry_transfer_input.lower() == 'y':
        print("이체를 다시 진행합니다...")
        
        # 다계좌이체진행 자동/수동 여부 묻기
        retry_auto_transfer_input = input("🟠다계좌이체진행(자동)을 함께 진행할까요? (y/n): ")
        retry_auto_transfer = retry_auto_transfer_input.lower() == 'y'
        
        # 웹페이지 상태 확인 및 다계좌이체 페이지로 이동
        try:
            # 프레임 전환
            driver.switch_to.default_content()  # 기본 프레임으로 돌아가기
            frame_id = "hanaMainframe"
            driver.switch_to.frame(frame_id)
            
            # 이체 메뉴 클릭
            transfer_link = driver.find_element(By.XPATH, "//a[@title='이체' and text()='이체']")
            transfer_link.click()
            print("'이체' 메뉴 링크를 클릭했습니다.")
            tm.sleep(1)     
            
            # 다계좌 이체 버튼 클릭
            try:
                # 먼저 이미지 인식으로 다계좌이체 버튼 찾기 시도
                if locate_and_click("multisend"):
                    print("이미지 인식으로 다계좌이체 버튼 클릭 성공")
                else:
                    # 이미지 인식 실패 시 탭으로 이동
                    print("이미지 인식 실패, 탭으로 다계좌이체 버튼 찾기")
                    for _ in range(6):   # 탭 6번
                        pyautogui.press('tab')
                    pyautogui.press('enter')   
                    print("탭으로 다계좌이체 버튼 클릭 완료")
            except Exception as e:
                print(f"다계좌이체 버튼 클릭 중 오류: {e}")
                # 실패 시 기존 방식으로 시도
                for _ in range(6):   # 탭 6번
                    pyautogui.press('tab')
                pyautogui.press('enter')   
            tm.sleep(3)
            
            # 스크롤 조정
            scroll_amount = 1000
            scroll_down(scroll_amount)
            tm.sleep(1)
            scroll_up(scroll_amount // 2)
            
        except Exception as e:
            print(f"웹페이지 설정 중 오류 발생: {e}")
            print("수동으로 다계좌이체 페이지로 이동해주세요.")
        
        # 이체 정보 입력 - 현재 첫 번째 시트에서 데이터 다시 로드
        df_retry2 = pd.read_excel(excel_path, sheet_name=0)  # 현재 첫 번째 시트 사용
        processed_data_retry2 = []
        
        for index, row in df_retry2.iterrows():
            product_name = row.iloc[3]  # 제품명
            customer_name = row.iloc[4]  # 이름
            account_info = row.iloc[7]  # 은행+계좌번호
            amount = row.iloc[9]  # 금액

            # 계좌정보 전처리
            bank_name, account_number = preprocess_account_info(account_info)
            name_product = f"{customer_name}{product_name}"
            processed_data_retry2.append((bank_name, account_number, name_product, product_name, amount))
        
        for index, data in enumerate(processed_data_retry2):
            if index >= 10:
                break
            input_transfer_info(data, index)
            tm.sleep(0.5)  # 각 세트 완료 후 잠시 대기

        # 비밀번호 입력
        enter_password()

        # 자동/수동에 따른 처리
        if retry_auto_transfer:
            # 다계좌이체진행 버튼 클릭
            click_transfer_button()

            # 보이스피싱 예방 팝업 처리
            handle_voice_phishing_popup()
            print("이체가 완료되었습니다.")
        else:
            print("이체가 취소되었습니다. 필요시 수동으로 다계좌이체진행 버튼을 클릭하세요.")
        
        # 이체 완료 후 다시 시트 순서 변경 여부 묻기
        print("\n" + "="*50)
        print("이체 작업이 완료되었습니다!")
        print("="*50)
        
        # 재귀적으로 시트 순서 변경 여부 묻기
        sheet_organize_input = input("🟢시트 순서를 변경하시겠습니까? (y/n): ")
        if sheet_organize_input.lower() == 'y':
            print("시트 순서 변경을 시작합니다...")
            organize_excel_sheets(str(file_path))
            print("시트 순서 변경이 완료되었습니다.")
            
            # 다시 이체 진행 여부 묻기
            retry_transfer_input2 = input("🟢시트를 변경했으니 다시 이체를 진행할까요? (y/n): ")
            if retry_transfer_input2.lower() == 'y':
                print("이체를 다시 진행합니다...")
                
                # 다계좌이체진행 자동/수동 여부 묻기
                retry_auto_transfer_input2 = input("🟠다계좌이체진행(자동)을 함께 진행할까요? (y/n): ")
                retry_auto_transfer2 = retry_auto_transfer_input2.lower() == 'y'
                
                # 웹페이지 상태 확인 및 다계좌이체 페이지로 이동
                try:
                    # 프레임 전환
                    driver.switch_to.default_content()  # 기본 프레임으로 돌아가기
                    frame_id = "hanaMainframe"
                    driver.switch_to.frame(frame_id)
                    
                    # 이체 메뉴 클릭
                    transfer_link = driver.find_element(By.XPATH, "//a[@title='이체' and text()='이체']")
                    transfer_link.click()
                    print("'이체' 메뉴 링크를 클릭했습니다.")
                    tm.sleep(1)
                    
                    # 다계좌 이체 버튼 클릭
                    try:
                        # 먼저 이미지 인식으로 다계좌이체 버튼 찾기 시도
                        if locate_and_click("multisend"):
                            print("이미지 인식으로 다계좌이체 버튼 클릭 성공")
                        else:
                            # 이미지 인식 실패 시 탭으로 이동
                            print("이미지 인식 실패, 탭으로 다계좌이체 버튼 찾기")
                            for _ in range(6):   # 탭 6번
                                pyautogui.press('tab')
                            pyautogui.press('enter')   
                            print("탭으로 다계좌이체 버튼 클릭 완료")
                    except Exception as e:
                        print(f"다계좌이체 버튼 클릭 중 오류: {e}")
                        # 실패 시 기존 방식으로 시도
                        for _ in range(6):   # 탭 6번
                            pyautogui.press('tab')
                        pyautogui.press('enter')   
                    tm.sleep(3)
                    
                    # 스크롤 조정
                    scroll_amount = 1000
                    scroll_down(scroll_amount)
                    tm.sleep(1)
                    scroll_up(scroll_amount // 2)
                    
                except Exception as e:
                    print(f"웹페이지 설정 중 오류 발생: {e}")
                    print("수동으로 다계좌이체 페이지로 이동해주세요.")
                
                # 이체 정보 입력 - 현재 첫 번째 시트에서 데이터 다시 로드
                df_retry = pd.read_excel(excel_path, sheet_name=0)  # 현재 첫 번째 시트 사용
                processed_data_retry = []
                
                for index, row in df_retry.iterrows():
                    product_name = row.iloc[3]  # 제품명
                    customer_name = row.iloc[4]  # 이름
                    account_info = row.iloc[7]  # 은행+계좌번호
                    amount = row.iloc[9]  # 금액

                    # 계좌정보 전처리
                    bank_name, account_number = preprocess_account_info(account_info)
                    name_product = f"{customer_name}{product_name}"
                    processed_data_retry.append((bank_name, account_number, name_product, product_name, amount))
                
                for index, data in enumerate(processed_data_retry):
                    if index >= 10:
                        break
                    input_transfer_info(data, index)
                    tm.sleep(0.5)  # 각 세트 완료 후 잠시 대기

                # 비밀번호 입력
                enter_password()

                # 자동/수동에 따른 처리
                if retry_auto_transfer2:
                    # 다계좌이체진행 버튼 클릭
                    click_transfer_button()

                    # 보이스피싱 예방 팝업 처리
                    handle_voice_phishing_popup()
                    print("이체가 완료되었습니다.")
                else:
                    print("이체가 취소되었습니다. 필요시 수동으로 다계좌이체진행 버튼을 클릭하세요.")
                
                # 이체 완료 후 다시 시트 순서 변경 여부 묻기
                print("\n" + "="*50)
                print("이체 작업이 완료되었습니다!")
                print("="*50)
                
                # 재귀적으로 시트 순서 변경 여부 묻기
                sheet_organize_input2 = input("🟢시트 순서를 변경하시겠습니까? (y/n): ")
                if sheet_organize_input2.lower() == 'y':
                    print("시트 순서 변경을 시작합니다...")
                    organize_excel_sheets(str(file_path))
                    print("시트 순서 변경이 완료되었습니다.")
                    
                    # 다시 이체 진행 여부 묻기
                    retry_transfer_input3 = input("🟢시트를 변경했으니 다시 이체를 진행할까요? (y/n): ")
                    if retry_transfer_input3.lower() == 'y':
                        print("이체를 다시 진행합니다...")
                        
                        # 다계좌이체진행 자동/수동 여부 묻기
                        retry_auto_transfer_input3 = input("🟠다계좌이체진행(자동)을 함께 진행할까요? (y/n): ")
                        retry_auto_transfer3 = retry_auto_transfer_input3.lower() == 'y'
                        
                        # 웹페이지 상태 확인 및 다계좌이체 페이지로 이동
                        try:
                            # 프레임 전환
                            driver.switch_to.default_content()  # 기본 프레임으로 돌아가기
                            frame_id = "hanaMainframe"
                            driver.switch_to.frame(frame_id)
                            
                            # 이체 메뉴 클릭
                            transfer_link = driver.find_element(By.XPATH, "//a[@title='이체' and text()='이체']")
                            transfer_link.click()
                            print("'이체' 메뉴 링크를 클릭했습니다.")
                            tm.sleep(1)
                            
                            # 다계좌 이체 버튼 클릭
                            try:
                                # 먼저 이미지 인식으로 다계좌이체 버튼 찾기 시도
                                if locate_and_click("multisend"):
                                    print("이미지 인식으로 다계좌이체 버튼 클릭 성공")
                                else:
                                    # 이미지 인식 실패 시 탭으로 이동
                                    print("이미지 인식 실패, 탭으로 다계좌이체 버튼 찾기")
                                    for _ in range(6):   # 탭 6번
                                        pyautogui.press('tab')
                                    pyautogui.press('enter')   
                                    print("탭으로 다계좌이체 버튼 클릭 완료")
                            except Exception as e:
                                print(f"다계좌이체 버튼 클릭 중 오류: {e}")
                                # 실패 시 기존 방식으로 시도
                                for _ in range(6):   # 탭 6번
                                    pyautogui.press('tab')
                                pyautogui.press('enter')   
                            tm.sleep(3)
                            
                            # 스크롤 조정
                            scroll_amount = 1000
                            scroll_down(scroll_amount)
                            tm.sleep(1)
                            scroll_up(scroll_amount // 2)
                            
                        except Exception as e:
                            print(f"웹페이지 설정 중 오류 발생: {e}")
                            print("수동으로 다계좌이체 페이지로 이동해주세요.")
                        
                        # 이체 정보 입력 - 현재 첫 번째 시트에서 데이터 다시 로드
                        df_retry3 = pd.read_excel(excel_path, sheet_name=0)  # 현재 첫 번째 시트 사용
                        processed_data_retry3 = []
                        
                        for index, row in df_retry3.iterrows():
                            product_name = row.iloc[3]  # 제품명
                            customer_name = row.iloc[4]  # 이름
                            account_info = row.iloc[7]  # 은행+계좌번호
                            amount = row.iloc[9]  # 금액

                            # 계좌정보 전처리
                            bank_name, account_number = preprocess_account_info(account_info)
                            name_product = f"{customer_name}{product_name}"
                            processed_data_retry3.append((bank_name, account_number, name_product, product_name, amount))
                        
                        for index, data in enumerate(processed_data_retry3):
                            if index >= 10:
                                break
                            input_transfer_info(data, index)
                            tm.sleep(0.5)  # 각 세트 완료 후 잠시 대기

                        # 비밀번호 입력
                        enter_password()

                        # 자동/수동에 따른 처리
                        if retry_auto_transfer3:
                            # 다계좌이체진행 버튼 클릭
                            click_transfer_button()

                            # 보이스피싱 예방 팝업 처리
                            handle_voice_phishing_popup()
                            print("이체가 완료되었습니다.")
                        else:
                            print("이체가 취소되었습니다. 필요시 수동으로 다계좌이체진행 버튼을 클릭하세요.")
                    else:
                        print("이체를 건너뜁니다.")
                else:
                    print("시트 순서 변경을 건너뜁니다.")
            else:
                print("이체를 건너뜁니다.")
        else:
            print("시트 순서 변경을 건너뜁니다.")
    else:
        print("이체를 건너뜁니다.")

## 1.2ver 완성